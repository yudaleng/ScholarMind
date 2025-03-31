import pandas as pd
import numpy as np
import logging
from typing import List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.config.config_manager import load_config

# 配置日志
logger = logging.getLogger(__name__)

class ExcelFormatter:
    """Excel格式化工具，处理保存到Excel的特殊格式"""
    
    def format_excel(self, df: pd.DataFrame, output_path: str, separate_sheets: bool = True, ai_fields: List[str] = None) -> bool:
        """
        格式化DataFrame并保存到Excel
        
        参数:
            df: 输入的DataFrame
            output_path: 输出Excel文件路径
            separate_sheets: 是否将不同数据源的结果放在不同工作表
            
        返回:
            是否成功保存
        """
        try:
            formatted_df = df.copy()
            
            # 处理特殊数据类型
            for col in formatted_df.columns:
                formatted_df[col] = formatted_df[col].apply(self._format_cell_value)
            
            # 确保删除publication_date和year字段，统一使用publication_year
            for field in ['publication_date', 'year']:
                if field in formatted_df.columns:
                    formatted_df.drop(field, axis=1, inplace=True)
            
            # 创建工作簿
            wb = Workbook()
            
            # 确保'source_type'列存在
            if 'source_type' not in formatted_df.columns:
                formatted_df['source_type'] = 'unknown'
            
            # 定义常见字段和特定数据源字段
            common_fields = [
                'title', 'full_authors', 'journal', 'publication_year',
                'abstract', 'doi', 'doi_link', 'keywords'
            ]
            
            # PubMed特有字段
            pubmed_fields = [
                'pmid', 'pubmed_link', 'PMID', 'DP', 'EDAT', 'MHDA', 'PHST', 'LID', 'MH',
                'OTO', 'OT', 'FAU', 'AU', 'AD', 'LA', 'PT', 'DEP', 'PL', 'TA', 'JT', 'mesh_terms'
            ]
            
            # WOS特有字段
            wos_fields = [
                'wos_id', 'wos_link', 'UT', 'CU', 'PY', 'VL', 'IS', 'BP', 'EP', 'DI', 'WC',
                'SC', 'GA', 'NR', 'TC', 'Z9', 'PU', 'PI', 'PA', 'AU_1', 'AF_1', 'C1', 'RP'
            ]
            
            # ScienceDirect特有字段
            sciencedirect_fields = [
                'url', 'sciencedirect_link', 'volume', 'issue', 'pages'
            ]
            
            # 从配置中获取期刊指标字段
            metric_cols = self._get_metric_columns(formatted_df)
            
            # 获取AI生成的字段
            ai_columns = ['summary']
            if 'summary' not in formatted_df.columns:
                ai_columns = []
                
            # 检查是否有其他AI生成的字段
            for col in ai_fields:
                ai_columns.append(col)
            
            # 1. 首先创建Results表 (合并后的结果)
            results_sheet = wb.active
            results_sheet.title = "Results"
            
            # Results表包含基本字段、期刊指标和AI字段
            results_columns = common_fields.copy()
            results_columns.extend(metric_cols)
            results_columns.extend(ai_columns)
            
            # 过滤只存在于DataFrame中的列
            results_columns = [col for col in results_columns if col in formatted_df.columns]
            
            # 确保source_type字段存在于Results表中
            if 'source_type' not in results_columns:
                results_columns.append('source_type')
                
            results_df = formatted_df[results_columns].copy()
            self._add_data_to_sheet(results_sheet, results_df)
            
            # 2. 创建WOS表（仅包含WOS来源的记录）
            wos_df = formatted_df[formatted_df['source_type'] == 'wos'].copy()
            if not wos_df.empty:
                wos_sheet = wb.create_sheet("WOS")
                
                if separate_sheets:
                    # 仅选择WOS特有字段和必要的基本字段
                    wos_columns = common_fields.copy()
                    wos_columns.extend([col for col in wos_fields if col in formatted_df.columns])
                else:
                    # 包含所有字段
                    wos_columns = [col for col in formatted_df.columns]
                
                # 确保wos_columns中的列在DataFrame中存在
                wos_columns = [col for col in wos_columns if col in wos_df.columns]
                wos_filtered_df = wos_df[wos_columns].copy()
                self._add_data_to_sheet(wos_sheet, wos_filtered_df)
            
            # 3. 创建PubMed表（仅包含PubMed来源的记录）
            pubmed_df = formatted_df[formatted_df['source_type'] == 'pubmed'].copy()
            if not pubmed_df.empty:
                pubmed_sheet = wb.create_sheet("PubMed")
                
                if separate_sheets:
                    # 仅选择PubMed特有字段和必要的基本字段
                    pubmed_columns = common_fields.copy()
                    pubmed_columns.extend([col for col in pubmed_fields if col in formatted_df.columns])
                else:
                    # 包含所有字段
                    pubmed_columns = [col for col in formatted_df.columns]
                
                # 确保pubmed_columns中的列在DataFrame中存在
                pubmed_columns = [col for col in pubmed_columns if col in pubmed_df.columns]
                pubmed_filtered_df = pubmed_df[pubmed_columns].copy()
                self._add_data_to_sheet(pubmed_sheet, pubmed_filtered_df)
            
            # 4. 创建ScienceDirect表（仅包含ScienceDirect来源的记录）
            sciencedirect_df = formatted_df[formatted_df['source_type'] == 'sciencedirect'].copy()
            if not sciencedirect_df.empty:
                sciencedirect_sheet = wb.create_sheet("ScienceDirect")
                
                if separate_sheets:
                    # 仅选择ScienceDirect特有字段和必要的基本字段
                    sciencedirect_columns = common_fields.copy()
                    sciencedirect_columns.extend([col for col in sciencedirect_fields if col in formatted_df.columns])
                else:
                    # 包含所有字段
                    sciencedirect_columns = [col for col in formatted_df.columns]
                
                # 确保sciencedirect_columns中的列在DataFrame中存在
                sciencedirect_columns = [col for col in sciencedirect_columns if col in sciencedirect_df.columns]
                sciencedirect_filtered_df = sciencedirect_df[sciencedirect_columns].copy()
                self._add_data_to_sheet(sciencedirect_sheet, sciencedirect_filtered_df)
            
            # 保存Excel文件
            wb.save(output_path)
            logger.info(f"Excel数据已成功保存到: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件时出错: {str(e)}")
            return False
    
    def _get_metric_columns(self, df: pd.DataFrame) -> List[str]:
        """
        从配置中获取期刊指标字段
        
        参数:
            df: 数据DataFrame，用于检查字段是否存在
            
        返回:
            期刊指标字段列表
        """
        try:
            # 加载配置
            config = load_config()
            journal_metrics_config = config.get("journal_metrics", {})
            
            # 获取要查询的指标和列名映射
            metrics_to_fetch = journal_metrics_config.get("metrics_to_fetch", [])
            metrics_column_mapping = journal_metrics_config.get("metrics_column_mapping", {})
            
            # 构建期刊指标字段列表
            metric_cols = []
            
            for metric in metrics_to_fetch:
                # 获取该指标对应的列名
                column_name = metrics_column_mapping.get(metric, metric)
                
                # 如果列名存在于DataFrame中，添加到指标字段列表
                if column_name in df.columns:
                    metric_cols.append(column_name)
            
            return metric_cols
        except Exception as e:
            logger.error(f"获取期刊指标字段时出错: {str(e)}")
            # 如果出错，退回到原始方法
            return [col for col in df.columns if col.startswith('jcr_') 
                    or col.startswith('sjr_') or col.startswith('cas_')]
    
    def _add_data_to_sheet(self, ws, df):
        """
        添加数据到工作表并设置格式
        
        参数:
            ws: 工作表对象
            df: 要添加的DataFrame数据
        """
        # 将DataFrame数据添加到工作表
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 应用表头样式
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置列宽
        column_width_mapping = {
            'title': 60,
            'abstract': 60,
            'ai_summary': 60,
            'authors': 30,
            'full_authors': 40,
            'journal': 25,
            'publication_year': 20,
            'pmid': 15,
            'doi': 15,
            'wos_id': 15,
            'affiliation': 40,
            'keywords': 30,
            'mesh_terms': 30,
            'publication_type': 20,
            'source_type': 15,
            'doi_link': 25,
            'pubmed_link': 25,
            'wos_link': 25,
            'sciencedirect_link': 25,
            'impact_factor': 15,
            '中科院分区': 15,
            'sciif': 15,
            'sci': 15,
            'sciUp': 15
        }
        
        # 设置所有列的宽度
        for col in range(1, ws.max_column + 1):
            column_letter = ws.cell(1, col).column_letter
            column_name = ws.cell(1, col).value
            
            # 根据列名设置宽度，如果列名不在映射中，设置为默认值20
            ws.column_dimensions[column_letter].width = column_width_mapping.get(column_name, 20)
        
        # 设置行高
        ws.row_dimensions[1].height = 25  # 表头行高
        
        # 设置数据行样式
        data_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 应用样式到所有单元格
        for row in range(1, ws.max_row + 1):
            row_height = 0
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell.border = thin_border
                if row > 1:  # 数据行
                    cell.alignment = data_alignment
                    # 根据内容调整行高
                    if cell.value and isinstance(cell.value, str):
                        # 每100个字符增加15单位高度
                        content_height = 15 + min(len(cell.value) // 100 * 15, 100)
                        row_height = max(row_height, content_height)
            
            # 设置行高，最小为20，最大为150
            if row > 1 and row_height > 0:
                ws.row_dimensions[row].height = min(max(row_height, 20), 150)
        
        # 添加超链接
        title_col = None
        pubmed_link_col = None
        wos_link_col = None
        doi_link_col = None
        sciencedirect_link_col = None
        
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header == 'title':
                title_col = col
            elif header == 'pubmed_link':
                pubmed_link_col = col
            elif header == 'wos_link':
                wos_link_col = col
            elif header == 'doi_link':
                doi_link_col = col
            elif header == 'sciencedirect_link':
                sciencedirect_link_col = col
        
        # 设置doi_link列的超链接
        if doi_link_col:
            for row in range(2, ws.max_row + 1):
                doi_link_cell = ws.cell(row, doi_link_col)
                if doi_link_cell.value:
                    # 设置doi_link的超链接
                    doi_link_cell.hyperlink = doi_link_cell.value
                    doi_link_cell.font = Font(color="0563C1", underline="single")
        
        # 添加标题超链接
        if title_col:
            for row in range(2, ws.max_row + 1):
                title_cell = ws.cell(row, title_col)
                link = None
                
                # 优先使用PubMed链接
                if pubmed_link_col and ws.cell(row, pubmed_link_col).value:
                    link = ws.cell(row, pubmed_link_col).value
                # 其次使用WOS链接
                elif wos_link_col and ws.cell(row, wos_link_col).value:
                    link = ws.cell(row, wos_link_col).value
                # 再次使用ScienceDirect链接
                elif sciencedirect_link_col and ws.cell(row, sciencedirect_link_col).value:
                    link = ws.cell(row, sciencedirect_link_col).value
                # 最后使用DOI链接
                elif doi_link_col and ws.cell(row, doi_link_col).value:
                    link = ws.cell(row, doi_link_col).value
                
                if link:
                    title_cell.hyperlink = link
                    title_cell.font = Font(color="0563C1", underline="single")
        
        # 冻结首行
        ws.freeze_panes = "A2"
    
    def _format_cell_value(self, value: Any) -> Any:
        """
        格式化单元格值以适应Excel
        
        参数:
            value: 单元格值
            
        返回:
            格式化后的值
        """
        # 处理None值和NaN值
        if value is None:
            return ""
        
        # 特别处理pandas的NA/NaN值
        try:
            if pd.isna(value):
                return ""
        except (TypeError, ValueError):
            # 如果value不能用于pd.isna判断，则继续处理
            pass
            
        # 处理numpy数组
        if isinstance(value, np.ndarray):
            try:
                # 如果是单元素数组，返回该元素
                if value.size == 1:
                    return self._format_cell_value(value.item())
                # 否则转换为列表处理
                return self._format_cell_value(value.tolist())
            except Exception as e:
                logger.warning(f"处理numpy数组时出错: {str(e)}")
                return str(value)
        
        # 处理pandas Series
        if hasattr(value, 'iloc') and hasattr(value, 'to_list'):
            try:
                return self._format_cell_value(value.to_list())
            except Exception as e:
                logger.warning(f"处理pandas Series时出错: {str(e)}")
                return str(value)
                
        # 处理列表
        if isinstance(value, list):
            # 空列表返回空字符串
            if not value:
                return ""
            
            # 尝试将列表中的每个元素转换为字符串，然后用逗号连接
            try:
                # 过滤掉None值并递归处理列表中的每个元素
                filtered_values = []
                for item in value:
                    if item is not None:
                        try:
                            if not pd.isna(item):
                                filtered_values.append(self._format_cell_value(item))
                        except (TypeError, ValueError):
                            filtered_values.append(self._format_cell_value(item))
                
                return ", ".join(str(v) for v in filtered_values if v != "")
            except Exception as e:
                logger.warning(f"转换列表值时出错: {str(e)}")
                # 如果转换出错，尝试返回第一个非None元素
                for item in value:
                    try:
                        if item is not None and not pd.isna(item):
                            return str(item)
                    except (TypeError, ValueError):
                        if item is not None:
                            return str(item)
                return ""
        
        # 处理字典
        if isinstance(value, dict):
            try:
                # 将字典转换为字符串表示
                return str(value)
            except Exception as e:
                logger.warning(f"转换字典值时出错: {str(e)}")
                return ""
        
        # 处理numpy类型
        if isinstance(value, (np.int64, np.int32, np.float64, np.float32)):
            return float(value) if isinstance(value, (np.float64, np.float32)) else int(value)
        
        # 其他类型尝试转换为字符串
        try:
            return str(value)
        except Exception as e:
            logger.warning(f"转换值时出错: {str(e)}")
            return "" 